import io
import json
import re
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from auth.dependencies import get_current_user, require_manager_or_above
from database import get_db
from sync.models import SyncedBatch, SyncedDpiRecord, SyncedSubjectScore

from .models import ExcelBatchRegistry, TraineeStreamReference
from .schemas import RowResult, ScoresUploadResult, StreamReferenceResponse

router = APIRouter(prefix="/scores", tags=["scores-upload"])

_DEFAULT_TEMPLATE_SUBJECTS = ["Java", "Python", "WebTech", "AIML", "Agile", "BizSkill"]

# Reserved (non-subject) columns, matched by header name — case/space/punctuation
# insensitive. Every other column with a header is treated as a subject score,
# named after its header text, so an Excel with more (or fewer, or reordered)
# subject columns than the default template is picked up automatically.
_RESERVED_COLUMNS = {
    "empid": "emp_id",
    "name": "name",
    "subbatch": "sub_batch",
    "dpi": "dpi",
    "stream": "stream",
}


def _normalize_header(value) -> str:
    return re.sub(r"[^a-z0-9]", "", str(value or "").lower())


def _read_column_layout(ws) -> tuple[dict[str, int], list[tuple[int, str]]]:
    """Map the header row to reserved-column indices and (index, subject_name)
    pairs for every other non-empty header. Raises HTTPException if a
    required reserved column is missing."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())

    reserved_idx: dict[str, int] = {}
    subject_columns: list[tuple[int, str]] = []
    for idx, header in enumerate(header_row):
        if header is None or str(header).strip() == "":
            continue
        key = _RESERVED_COLUMNS.get(_normalize_header(header))
        if key:
            reserved_idx[key] = idx
        else:
            subject_columns.append((idx, str(header).strip()))

    missing = [c for c in ("emp_id", "name", "sub_batch", "dpi") if c not in reserved_idx]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Excel is missing required column(s): {', '.join(missing)}",
        )

    return reserved_idx, subject_columns


# ── Template ──────────────────────────────────────────────────────────

@router.get("/excel-template")
def download_template(_=Depends(get_current_user)):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Employee Scores"

    headers = ["Emp Id", "Name", "Sub Batch", "DPI", "Stream", *_DEFAULT_TEMPLATE_SUBJECTS]
    hdr_fill = PatternFill(start_color="1E4D8C", end_color="1E4D8C", fill_type="solid")
    hdr_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")

    ws.append(["EMP001", "John Doe",   "A1", 3.5, "Java Development",  78.5, 65.0, 72.0, 55.0, 80.0, 70.0])
    ws.append(["EMP002", "Jane Smith", "A2", 4.2, "AI/ML Engineering", 60.0, 85.0, 70.0, 92.0, 75.0, 68.0])

    col_widths = [12, 20, 12, 8, 25, 10, 10, 12, 10, 10, 12]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=employee_scores_template.xlsx"},
    )


# ── Batch info / Excel-batch list ─────────────────────────────────────

@router.get("/batch-info/{batch_name}")
def batch_info(
    batch_name: str,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    dpi_count = (
        db.query(SyncedDpiRecord)
        .filter(SyncedDpiRecord.batch_name == batch_name)
        .count()
    )
    excel_reg = (
        db.query(ExcelBatchRegistry)
        .filter(ExcelBatchRegistry.batch_name == batch_name)
        .first()
    )
    return {
        "batch_name": batch_name,
        "dpi_count": dpi_count,
        "has_existing": dpi_count > 0,
        "excel_managed": excel_reg is not None,
        "uploaded_at": excel_reg.uploaded_at if excel_reg else None,
    }


@router.get("/excel-batches")
def list_excel_batches(db: Session = Depends(get_db), _=Depends(get_current_user)):
    rows = db.query(ExcelBatchRegistry).all()
    return [{"batch_name": r.batch_name, "uploaded_at": r.uploaded_at, "trainee_count": r.trainee_count} for r in rows]


# ── Upload ────────────────────────────────────────────────────────────

@router.post("/upload-excel", response_model=ScoresUploadResult)
async def upload_excel(
    batch_name: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _=Depends(require_manager_or_above),
):
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content))
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid or corrupt Excel file")

    ws = wb.active
    reserved_idx, subject_columns = _read_column_layout(ws)
    subject_names = [name for _, name in subject_columns]
    row_results: list[RowResult] = []
    succeeded = 0
    failed = 0
    now = datetime.now(timezone.utc).isoformat()

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not any(cell is not None for cell in row):
                continue

            def _get(i: int, r=row):
                return r[i] if len(r) > i else None

            emp_id     = _get(reserved_idx["emp_id"])
            name       = _get(reserved_idx["name"])
            sub_batch  = _get(reserved_idx["sub_batch"])
            dpi_val    = _get(reserved_idx["dpi"])
            stream_val = _get(reserved_idx["stream"]) if "stream" in reserved_idx else None

            if not emp_id or not name:
                row_results.append(RowResult(
                    row=row_idx, trainee_id=str(emp_id or ""),
                    status="error", detail="Emp Id and Name are required",
                ))
                failed += 1
                continue

            emp_id_str    = str(emp_id).strip()
            name_str      = str(name).strip()
            sub_batch_str = str(sub_batch).strip() if sub_batch else None

            try:
                dpi_float = float(dpi_val)
                if not (0.0 <= dpi_float <= 5.0):
                    raise ValueError
            except (TypeError, ValueError):
                row_results.append(RowResult(
                    row=row_idx, trainee_id=emp_id_str,
                    status="error", detail="DPI must be a number between 0 and 5",
                ))
                failed += 1
                continue

            subject_pairs = [(subj_name, _get(idx)) for idx, subj_name in subject_columns]
            subject_scores: list[tuple[str, float]] = []
            row_error: Optional[str] = None
            for subj_name, subj_val in subject_pairs:
                if subj_val is None:
                    continue
                try:
                    s = float(subj_val)
                    if not (0.0 <= s <= 100.0):
                        row_error = f"{subj_name} score must be between 0 and 100"
                        break
                    subject_scores.append((subj_name, s))
                except (TypeError, ValueError):
                    row_error = f"{subj_name} score must be a number"
                    break

            if row_error:
                row_results.append(RowResult(row=row_idx, trainee_id=emp_id_str, status="error", detail=row_error))
                failed += 1
                continue

            # ── Write directly to synced tables (Excel has priority) ──

            existing_dpi = (
                db.query(SyncedDpiRecord)
                .filter(SyncedDpiRecord.trainee_id == emp_id_str)
                .first()
            )
            if existing_dpi:
                existing_dpi.batch_name = batch_name
                existing_dpi.trainee_name = name_str
                existing_dpi.dpi = dpi_float
                existing_dpi.sub_batch = sub_batch_str
                existing_dpi.synced_at = now
            else:
                db.add(SyncedDpiRecord(
                    trainee_id=emp_id_str,
                    batch_name=batch_name,
                    trainee_name=name_str,
                    dpi=dpi_float,
                    location=None,
                    sub_batch=sub_batch_str,
                    synced_at=now,
                ))

            # Delete existing synced scores for this trainee+batch, then re-insert
            db.query(SyncedSubjectScore).filter(
                SyncedSubjectScore.trainee_id == emp_id_str,
                SyncedSubjectScore.batch_name == batch_name,
            ).delete(synchronize_session="fetch")
            for subj_name, score_val in subject_scores:
                db.add(SyncedSubjectScore(
                    external_id=f"excel-{emp_id_str}-{subj_name}",
                    batch_name=batch_name,
                    trainee_id=emp_id_str,
                    trainee_name=name_str,
                    subject_name=subj_name,
                    subject_id=None,
                    exam_name="Excel Upload",
                    score=score_val,
                    synced_at=now,
                ))

            # ── Stream reference ──
            if stream_val:
                stream_str = str(stream_val).strip()
                if stream_str:
                    ref = (
                        db.query(TraineeStreamReference)
                        .filter(TraineeStreamReference.trainee_id == emp_id_str)
                        .first()
                    )
                    if ref:
                        ref.batch_name = batch_name
                        ref.stream_name = stream_str
                        ref.updated_at = now
                    else:
                        db.add(TraineeStreamReference(
                            trainee_id=emp_id_str,
                            batch_name=batch_name,
                            stream_name=stream_str,
                            updated_at=now,
                        ))

            row_results.append(RowResult(row=row_idx, trainee_id=emp_id_str, status="ok"))
            succeeded += 1

    # Upsert synced_batches entry so the batch appears everywhere
    existing_batch = (
        db.query(SyncedBatch).filter(SyncedBatch.batch_name == batch_name).first()
    )
    if existing_batch:
        existing_batch.subjects_json = json.dumps(subject_names)
        if succeeded > existing_batch.trainee_count:
            existing_batch.trainee_count = succeeded
        existing_batch.synced_at = now
    else:
        db.add(SyncedBatch(
            batch_name=batch_name,
            subjects_json=json.dumps(subject_names),
            trainee_count=succeeded,
            synced_at=now,
        ))

    # Register / update Excel batch registry
    reg = (
        db.query(ExcelBatchRegistry)
        .filter(ExcelBatchRegistry.batch_name == batch_name)
        .first()
    )
    if reg:
        reg.uploaded_at = now
        reg.trainee_count = succeeded
    else:
        db.add(ExcelBatchRegistry(
            batch_name=batch_name,
            uploaded_at=now,
            trainee_count=succeeded,
        ))

    db.commit()

    return ScoresUploadResult(
        rows_processed=succeeded + failed,
        rows_succeeded=succeeded,
        rows_failed=failed,
        row_results=row_results,
        sync_triggered=False,  # direct write — no sync step needed
    )


# ── Stream references ─────────────────────────────────────────────────

@router.get("/stream-references", response_model=list[StreamReferenceResponse])
def list_stream_references(
    batch_name: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    q = db.query(TraineeStreamReference)
    if batch_name:
        q = q.filter(TraineeStreamReference.batch_name == batch_name)
    return q.all()
